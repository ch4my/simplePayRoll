from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
import database

def export_to_excel(user_data: dict, filename: str = None):
    """
    Export payroll data to Excel with metadata header.
    Returns the filename used.
    """
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"payroll_export_{timestamp}.xlsx"
    
    # Fetch all records
    records = database.fetch_all()
    
    # Create workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Payroll Data"
    
    # --- Metadata Header Section ---
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True, size=12)
    
    ws.merge_cells('A1:K1')
    ws['A1'] = "PAYROLL SYSTEM - EMPLOYEE DATA EXPORT"
    ws['A1'].font = Font(bold=True, size=14, color="2C3E50")
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    
    # Export metadata
    metadata = [
        ["Export Date:", datetime.now().strftime("%Y-%m-%d %H:%M:%S")],
        ["Exported By:", user_data.get('full_name') or user_data['username']],
        ["Username:", user_data['username']],
        ["Email:", user_data.get('email', 'N/A')],
        ["Total Records:", len(records)],
        ["", ""]  # Empty row for spacing
    ]
    
    for idx, (label, value) in enumerate(metadata, start=2):
        ws[f'A{idx}'] = label
        ws[f'A{idx}'].font = Font(bold=True)
        ws[f'B{idx}'] = value
    
    # --- Table Header ---
    header_row = 8
    headers = ['ID', 'Name', 'Company ID', 'Age', 'Role', 'Department', 
               'Period', 'Months', 'Overall Pay', 'Overall Deductions', 'Loan', 'Total Salary']
    
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=header_row, column=col_idx, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
    
    # --- Data Rows ---
    row_idx = header_row + 1
    for record in records:
        # record: id, name, company_id, age, role, department, months, loan, 
        #         deduction, overall_salary, total_salary, created_at, start_month, end_month, currency
        
        monthly_deduction = int(record[8]) if record[8] is not None else 0
        monthly_net = int(record[9]) if record[9] is not None else 0
        monthly_gross = monthly_net + monthly_deduction
        
        months = int(record[6]) if record[6] is not None else 1
        loan = int(record[7]) if record[7] is not None else 0
        
        total_gross = monthly_gross * months
        total_deductions = monthly_deduction * months
        total_net = int(record[10]) if record[10] is not None else 0
        
        # Format period
        start_month = str(record[12] or '').strip()
        end_month = str(record[13] or '').strip()
        period = f"{start_month} to {end_month}" if start_month and end_month else "N/A"
        
        row_data = [
            record[0],  # ID
            record[1],  # Name
            record[2],  # Company ID
            record[3],  # Age
            record[4],  # Role
            record[5],  # Department
            period,     # Period
            months,     # Months
            f"PHP {total_gross:,}",  # Overall Pay
            f"PHP {total_deductions:,}",  # Overall Deductions
            f"PHP {loan:,}",  # Loan
            f"PHP {total_net:,}"  # Total Salary
        ]
        
        for col_idx, value in enumerate(row_data, start=1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            if col_idx in [9, 10, 11, 12]:  # Money columns
                cell.alignment = Alignment(horizontal='right')
        
        row_idx += 1
    
    # Adjust column widths
    column_widths = [8, 20, 15, 8, 15, 15, 20, 10, 18, 20, 15, 18]
    for idx, width in enumerate(column_widths, start=1):
        ws.column_dimensions[chr(64 + idx)].width = width
    
    # Save workbook
    wb.save(filename)
    return filename


def export_to_pdf(user_data: dict, filename: str = None):
    """
    Export payroll data to PDF with metadata header.
    Returns the filename used.
    """
    if filename is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        filename = f"payroll_export_{timestamp}.pdf"
    
    # Fetch all records
    records = database.fetch_all()
    
    # Create PDF document
    doc = SimpleDocTemplate(filename, pagesize=landscape(letter),
                           rightMargin=30, leftMargin=30,
                           topMargin=30, bottomMargin=18)
    
    # Container for PDF elements
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        textColor=colors.HexColor('#2C3E50'),
        spaceAfter=12,
        alignment=1  # Center
    )
    
    # --- Title ---
    title = Paragraph("PAYROLL SYSTEM - EMPLOYEE DATA EXPORT", title_style)
    elements.append(title)
    elements.append(Spacer(1, 0.2*inch))
    
    # --- Metadata Table ---
    export_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    metadata_data = [
        ['Export Date:', export_date],
        ['Exported By:', user_data.get('full_name') or user_data['username']],
        ['Username:', user_data['username']],
        ['Email:', user_data.get('email', 'N/A')],
        ['Total Records:', str(len(records))]
    ]
    
    metadata_table = Table(metadata_data, colWidths=[1.5*inch, 4*inch])
    metadata_table.setStyle(TableStyle([
        ('FONT', (0, 0), (0, -1), 'Helvetica-Bold', 10),
        ('FONT', (1, 0), (1, -1), 'Helvetica', 10),
        ('TEXTCOLOR', (0, 0), (-1, -1), colors.HexColor('#2C3E50')),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    
    elements.append(metadata_table)
    elements.append(Spacer(1, 0.3*inch))
    
    # --- Data Table ---
    table_data = [['ID', 'Name', 'Company ID', 'Age', 'Role', 'Department', 
                   'Period', 'Overall Pay', 'Overall Ded.', 'Loan', 'Total Salary']]
    
    for record in records:
        monthly_deduction = int(record[8]) if record[8] is not None else 0
        monthly_net = int(record[9]) if record[9] is not None else 0
        monthly_gross = monthly_net + monthly_deduction
        
        months = int(record[6]) if record[6] is not None else 1
        loan = int(record[7]) if record[7] is not None else 0
        
        total_gross = monthly_gross * months
        total_deductions = monthly_deduction * months
        total_net = int(record[10]) if record[10] is not None else 0
        
        # Format period
        start_month = str(record[12] or '').strip()
        end_month = str(record[13] or '').strip()
        period = f"{start_month} - {end_month}" if start_month and end_month else "N/A"
        
        table_data.append([
            str(record[0]),
            str(record[1])[:15],  # Truncate long names
            str(record[2]),
            str(record[3]),
            str(record[4])[:12],
            str(record[5])[:12],
            period,
            f"₱{total_gross:,}",
            f"₱{total_deductions:,}",
            f"₱{loan:,}",
            f"₱{total_net:,}"
        ])
    
    # Create table with appropriate column widths
    col_widths = [0.4*inch, 1*inch, 0.9*inch, 0.4*inch, 0.9*inch, 
                  0.9*inch, 1.2*inch, 1*inch, 1*inch, 0.8*inch, 1*inch]
    
    data_table = Table(table_data, colWidths=col_widths, repeatRows=1)
    data_table.setStyle(TableStyle([
        # Header row styling
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        
        # Data rows styling
        ('TEXTCOLOR', (0, 1), (-1, -1), colors.HexColor('#2C3E50')),
        ('ALIGN', (0, 1), (0, -1), 'CENTER'),  # ID column
        ('ALIGN', (3, 1), (3, -1), 'CENTER'),  # Age column
        ('ALIGN', (7, 1), (-1, -1), 'RIGHT'),  # Money columns
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
        
        # Grid
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 5),
    ]))
    
    elements.append(data_table)
    
    # Build PDF
    doc.build(elements)
    return filename